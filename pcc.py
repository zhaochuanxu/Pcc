import numpy as np
import pandas as pd

import openpyxl


from openpyxl import load_workbook

#导入文件获取workbook
wb = load_workbook('expression.xlsx')

#获取worksheet
ws = wb['Sheet1']

#关闭sheet
wb.close()

print(ws)

print(ws.cell(row=1,column=3).value)


rows=ws.max_row   #获取行数
cols=ws.max_column    #获取列数
for i in range(2,rows+1)
    x=pd.Series()









#打开文件：
#from openpyxl import load_workbook
#excel=load_workbook('E:/test.xlsx')
#获取sheet：
#table = excel.get_sheet_by_name('Sheet1')   #通过表名获取
#获取行数和列数：
#rows=table.max_row   #获取行数
#cols=table.max_column    #获取列数
#获取单元格值：
#Data=table.cell(row=row,column=col).value  #获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value